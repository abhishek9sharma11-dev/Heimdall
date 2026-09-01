#!/usr/bin/env python3
"""Day ops: auto-start webinar stacks 30 min before start; track peak attendees;
deliver a completed-session report to Slack, then remove transient session state.

Reads schedules/today_sessions.json. For each enabled session:
  - At (session_start - start_lead_minutes): start bridge+python via hermes_slots.sh
  - For first peak_window_minutes after session_start: poll participant counts every 10s
  - At first payment-link drop: snapshot attendees (retention)
  - At session_end + 5 min: deliver the report to Slack and clean up

Does not stop other sessions. Safe to leave running all day.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import time
import sys
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

# The launcher invokes this file by absolute path. Ensure repository packages
# remain importable in that mode as well as when run from the repository root.
_REPO_ROOT = str(Path(__file__).resolve().parents[1])
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

try:
    from scripts.slack_reporting import finalize_report, metric_row
except ModuleNotFoundError:
    from slack_reporting import finalize_report, metric_row

from dashboard.session_time import normalize_session_time

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "schedules" / "today_sessions.json"
HERMES = ROOT / "scripts" / "hermes_slots.sh"
PEAK_OUT = Path("/tmp/hermes-peak-attendees.json")
PAY_DROP_OUT = Path("/tmp/hermes-payment-drop-attendees.json")
SHEET_SYNC_OUT = Path("/tmp/hermes-sheet-sync.json")
LOG = Path("/tmp/day-ops-runner.log")
TZ = ZoneInfo("Asia/Kolkata")
TICK_SEC = 5
PEAK_INTERVAL = 10
DAILY_MD_INTERVAL = 60
# Wait a few minutes after session_end so late payments can land before reporting.
SHEET_SYNC_GRACE_MIN = 5
SLACK_REPORT_OUT = Path("/tmp/hermes-slack-reports.json")
PAYMENT_LINK_RE = re.compile(r"https?://link\.outskill\.com/[^\s<>\"']+", re.I)

EVAL = r"""
var text = (document.body && document.body.innerText) || '';
var n = null, attendees = null, participants = null;
var btns = Array.prototype.slice.call(document.querySelectorAll('button'));
for (var i = 0; i < btns.length; i++) {
  var a = (btns[i].getAttribute('aria-label') || '') + ' ' + (btns[i].innerText || '');
  var m = a.match(/(\d+)\s*participant/i) || a.match(/\[(\d+)\]/);
  if (m) { n = parseInt(m[1], 10); break; }
}
if (n === null) {
  var m2 = text.match(/(\d+)\s*Participants?/i);
  if (m2) n = parseInt(m2[1], 10);
}
var ma = text.match(/Attendees\s*\((\d+)\)/i);
if (ma) attendees = parseInt(ma[1], 10);
var mp = text.match(/Participants\s*\((\d+)\)/i);
if (mp) participants = parseInt(mp[1], 10);
return {footer: n, attendees: attendees, participants: participants};
"""


def log(msg: str) -> None:
    # stdout only — launcher redirects to LOG; avoid double-write
    print(f"{datetime.now(TZ).strftime('%H:%M:%S')} {msg}", flush=True)


def now() -> datetime:
    return datetime.now(TZ)


def parse_hhmmss(s: str) -> datetime:
    h, m, sec = map(int, normalize_session_time(s, allow_empty=False).split(":"))
    n = now()
    return n.replace(hour=h, minute=m, second=sec, microsecond=0)


def _valid_session(entry: object) -> dict:
    if not isinstance(entry, dict):
        raise ValueError("registration is not an object")
    sid = str(entry.get("id") or entry.get("meeting_id") or "").strip()
    if not sid:
        raise ValueError("registration has no id")
    if not entry.get("env_file"):
        raise ValueError("missing env_file")
    if not entry.get("schedule_file"):
        raise ValueError("missing schedule_file")
    int(entry.get("port"))
    normalize_session_time(entry.get("session_start_ist"), allow_empty=False)
    # Persistent worker registrations intentionally have no end time.
    if not entry.get("keep_connected"):
        normalize_session_time(entry.get("session_end_ist"), allow_empty=False)
    return entry


def load_manifest_sessions() -> list[dict]:
    """Load valid registrations and keep malformed entries from killing the worker."""
    if not MANIFEST.exists():
        log(f"manifest unavailable path={MANIFEST} — waiting for registrations")
        return []
    try:
        raw = json.loads(MANIFEST.read_text()).get("sessions") or []
    except Exception as exc:
        log(f"manifest discovery failed path={MANIFEST}: {type(exc).__name__}: {exc}")
        return []
    sessions: list[dict] = []
    for entry in raw:
        try:
            sessions.append(_valid_session(entry))
        except Exception as exc:
            sid = entry.get("id") if isinstance(entry, dict) else "<unknown>"
            log(
                f"discovery skipped session {sid} path={MANIFEST}: "
                f"{type(exc).__name__}: {exc}"
            )
    return sessions


def _env_values(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for line in path.read_text().splitlines():
        if "=" not in line or line.lstrip().startswith("#"):
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("'\"")
    return values


def health(port: int) -> dict:
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/health", timeout=3) as r:
            return json.loads(r.read().decode())
    except Exception:
        return {}


def eval_counts(port: int) -> dict:
    try:
        req = urllib.request.Request(
            f"http://127.0.0.1:{port}/eval",
            data=json.dumps({"code": EVAL}).encode(),
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read().decode()).get("result") or {}
    except Exception as e:
        return {"error": str(e)}


def load_peak() -> dict:
    if PEAK_OUT.exists():
        try:
            return json.loads(PEAK_OUT.read_text())
        except Exception:
            pass
    return {"updated_at": None, "slots": {}}


def save_peak(data: dict) -> None:
    PEAK_OUT.write_text(json.dumps(data, indent=2))


def load_pay_drops() -> dict:
    if PAY_DROP_OUT.exists():
        try:
            return json.loads(PAY_DROP_OUT.read_text())
        except Exception:
            pass
    return {
        "tag": "payment_link_drop_attendees_count",
        "updated_at": None,
        "sessions": {},
    }


def save_pay_drops(data: dict) -> None:
    PAY_DROP_OUT.write_text(json.dumps(data, indent=2) + "\n")


def load_slack_reports() -> dict:
    if SLACK_REPORT_OUT.exists():
        try:
            return json.loads(SLACK_REPORT_OUT.read_text())
        except Exception:
            pass
    return {"sessions": {}}


def save_slack_reports(data: dict) -> None:
    SLACK_REPORT_OUT.write_text(json.dumps(data, indent=2) + "\n")


def first_payment_from_schedule(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        doc = json.loads(path.read_text())
    except Exception:
        return None
    rows = []
    for it in doc.get("items") or []:
        text = it.get("text") or ""
        m = PAYMENT_LINK_RE.search(text)
        if not m:
            continue
        t = it.get("time") or ""
        if t:
            rows.append((t, text, m.group(0).rstrip(").,;")))
    if not rows:
        return None
    rows.sort(key=lambda x: x[0])
    t, text, url = rows[0]
    return {"time": t, "text": text, "url": url}


def best_count(counts: dict) -> int | None:
    for k in ("attendees", "footer", "participants"):
        v = counts.get(k)
        if isinstance(v, int) and v >= 0:
            return v
    return None


def capture_payment_drop(s: dict, first: dict, counts: dict, pay: dict) -> None:
    sid = s["id"]
    n = now()
    day = n.strftime("%Y-%m-%d")
    # One retention snapshot per session per calendar day — never overwrite older days.
    key = f"{day}:{sid}"
    sessions = pay.setdefault("sessions", {})
    if key in sessions and sessions[key].get("attendees_count") is not None:
        return
    # Clear legacy undated key only if it belongs to a previous day.
    legacy = sessions.get(sid)
    if legacy and not str(legacy.get("dropped_at") or "").startswith(day):
        pass  # leave historical legacy entry; prefer dated keys going forward
    entry = {
        "tag": "payment_link_drop_attendees_count",
        "session_id": sid,
        "meeting_id": s.get("meeting_id"),
        "port": s.get("port"),
        "date": day,
        "schedule_time": first["time"],
        "dropped_at": n.isoformat(),
        "payment_url": first.get("url") or "",
        "message_preview": (first.get("text") or "")[:160].replace("\n", " "),
        "attendees_count": best_count(counts),
        "footer": counts.get("footer"),
        "attendees": counts.get("attendees"),
        "participants": counts.get("participants"),
        "source": "day_ops",
    }
    sessions[key] = entry
    pay["updated_at"] = n.isoformat()
    save_pay_drops(pay)
    log(
        f"payment_link_drop_attendees_count {sid} :{s['port']} "
        f"at={first['time']} count={entry['attendees_count']} counts={counts}"
    )


def env_datetime(env_file: str, key: str) -> datetime | None:
    """Read a full ISO datetime (e.g. WEBINAR_END_AT) from a session's env file.

    Unlike session_start_ist/session_end_ist (time-of-day only, always assumed to be
    today), this carries the actual registration date, so it is what tells a session
    registered yesterday apart from one registered today.
    """
    env_path = ROOT / env_file
    if not env_path.exists():
        return None
    for line in env_path.read_text().splitlines():
        if line.startswith(f"{key}="):
            val = line.split("=", 1)[1].strip().strip("'\"")
            if not val:
                return None
            try:
                dt = datetime.fromisoformat(val)
            except ValueError:
                return None
            return dt if dt.tzinfo else dt.replace(tzinfo=TZ)
    return None


def persist_force_disable(session_id: str) -> None:
    """Write force_disabled=true back to the manifest so an expired session stays
    disabled across restarts instead of being auto-adopted again on a future day."""
    try:
        manifest = json.loads(MANIFEST.read_text())
        for entry in manifest.get("sessions") or []:
            if entry.get("id") == session_id:
                entry["force_disabled"] = True
                entry["enabled"] = False
        MANIFEST.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    except Exception as e:
        log(f"WARN could not persist force_disabled for {session_id}: {e}")


def forget_session(session: dict, report_files: list[str]) -> None:
    """Delete transient runtime state after a successful Slack delivery."""
    session_id = str(session.get("id") or "")
    meeting_id = str(session.get("meeting_id") or "")
    identifiers = {session_id, meeting_id}

    for raw in report_files:
        try:
            Path(raw).unlink(missing_ok=True)
        except OSError:
            pass

    for path, collection in (
        (PEAK_OUT, "slots"),
        (PAY_DROP_OUT, "sessions"),
        (SHEET_SYNC_OUT, "sessions"),
    ):
        try:
            data = json.loads(path.read_text()) if path.exists() else {}
            values = data.get(collection) or {}
            data[collection] = {
                key: value for key, value in values.items()
                if str(value.get("id") or value.get("session_id") or "") not in identifiers
                and str(value.get("meeting_id") or "") != meeting_id
            }
            path.write_text(json.dumps(data, indent=2) + "\n")
        except (OSError, json.JSONDecodeError):
            pass

    try:
        manifest = json.loads(MANIFEST.read_text()) if MANIFEST.exists() else {}
        manifest["sessions"] = [
            value for value in manifest.get("sessions") or []
            if str(value.get("id") or value.get("meeting_id") or "") not in identifiers
        ]
        MANIFEST.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")
    except (OSError, json.JSONDecodeError):
        pass

    for raw in (session.get("env_file"), session.get("schedule_file")):
        if not raw:
            continue
        path = ROOT / str(raw)
        if path.name == MANIFEST.name:
            continue
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
    log(f"forgot transient session {session_id}")


def start_slot(port: int, env_file: str) -> bool:
    h = health(port)
    state = h.get("meeting_state")
    if state in ("waiting", "in_meeting", "joining"):
        log(f"skip start :{port} already {state}")
        return True
    if h.get("status") == "ok":
        # bridge up but idle — stop then restart clean
        subprocess.run([str(HERMES), "stop", str(port)], cwd=str(ROOT), check=False)
        time.sleep(1)
    env_path = ROOT / env_file
    if not env_path.exists():
        log(f"ERROR missing env {env_file}")
        return False
    # require join URL
    text = env_path.read_text()
    if "MEETING_JOIN_URL=\n" in text or "MEETING_JOIN_URL=\r" in text or "MEETING_JOIN_URL=$" in text:
        log(f"ERROR :{port} empty MEETING_JOIN_URL in {env_file}")
        return False
    if "MEETING_JOIN_URL=" not in text:
        log(f"ERROR :{port} no MEETING_JOIN_URL in {env_file}")
        return False
    # empty value after =
    for line in text.splitlines():
        if line.startswith("MEETING_JOIN_URL="):
            val = line.split("=", 1)[1].strip().strip("'\"")
            if not val:
                log(f"ERROR :{port} empty join URL — cannot start")
                return False
    log(f"starting :{port} with {env_file}")
    r = subprocess.run(
        [str(HERMES), "start", str(port), env_file],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
    )
    if r.returncode != 0:
        log(f"start FAILED :{port}: {r.stdout[-400:]} {r.stderr[-400:]}")
        return False
    log(f"start OK :{port}: {r.stdout.strip().splitlines()[-1] if r.stdout.strip() else 'ok'}")
    return True


def try_fill_missing_join_from_excel(sessions: list[dict]) -> None:
    """If Claude 199$ (or any blocked slot) gains a zoom URL in Workshops (2).xlsx, write it into .env."""
    xlsx = Path("/Users/growthschool/Downloads/Workshops (2).xlsx")
    if not xlsx.exists():
        return
    try:
        import openpyxl
        import re
        from urllib.parse import urlparse, parse_qs
    except Exception:
        return
    try:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as e:
        log(f"excel reload skip: {e}")
        return
    urls_by_mid: dict[str, str] = {}
    for sheet in wb.sheetnames:
        mid = None
        url = None
        for row in wb[sheet].iter_rows(values_only=True):
            a = row[0] if row else None
            b = row[1] if row and len(row) > 1 else None
            a_s = "" if a is None else str(a).strip()
            b_s = "" if b is None else str(b).strip()
            if a_s.lower().startswith("webinar") and b_s:
                mid = re.sub(r"[^\d]", "", b_s.split(".")[0])
            if b_s.startswith("http") and "zoom.us/w/" in b_s:
                url = b_s.split()[0].strip()
        if mid and url:
            urls_by_mid[mid] = url
    for s in sessions:
        if s.get("enabled") and s.get("join_url_present"):
            continue
        mid = s["meeting_id"]
        url = urls_by_mid.get(mid)
        if not url:
            continue
        tk = parse_qs(urlparse(url).query).get("tk", [""])[0]
        env = ROOT / s["env_file"]
        text = env.read_text() if env.exists() else ""
        lines = []
        for line in text.splitlines():
            if line.startswith("MEETING_JOIN_URL="):
                lines.append(f"MEETING_JOIN_URL={url}")
            elif line.startswith("MEETING_WEBINAR_TOKEN="):
                lines.append(f"MEETING_WEBINAR_TOKEN={tk}")
            else:
                lines.append(line)
        if not any(l.startswith("MEETING_JOIN_URL=") for l in lines):
            lines.append(f"MEETING_JOIN_URL={url}")
            lines.append(f"MEETING_WEBINAR_TOKEN={tk}")
        env.write_text("\n".join(lines) + "\n")
        s["join_url_present"] = True
        if s.get("force_disabled"):
            s["enabled"] = False
            log(f"filled join URL for {s['id']} but force_disabled — leaving enabled=false")
        else:
            s["enabled"] = True
            log(f"filled join URL for {s['id']} mid={mid} from Workshops (2).xlsx")


def main() -> None:
    sessions = load_manifest_sessions()
    started: set[str] = set()
    last_peak_at: dict[str, float] = {}
    peak = load_peak()
    pay = load_pay_drops()
    slack_reports = load_slack_reports()
    first_pay: dict[str, dict] = {}
    last_daily_md = 0.0
    for s in sessions:
        fp = first_payment_from_schedule(ROOT / s["schedule_file"])
        if fp:
            first_pay[s["id"]] = fp
            log(
                f"  first_payment {s['id']} @ {fp['time']} → {fp['url'][:60]}"
            )
        else:
            log(f"  first_payment {s['id']} — none found in schedule")

    log(
        f"day_ops_runner live — {len(sessions)} sessions "
        f"({sum(1 for s in sessions if s.get('enabled'))} enabled)"
    )
    for s in sessions:
        lead = int(s.get("start_lead_minutes") or 30)
        st = parse_hhmmss(s["session_start_ist"])
        go = st - timedelta(minutes=lead)
        flag = "READY" if s.get("enabled") else "BLOCKED(no join URL)"
        log(
            f"  {s['id']} :{s['port']} start={s['session_start_ist']} "
            f"auto_start={go.strftime('%H:%M:%S')} {flag}"
        )

    # Mark already-live slots as started
    for s in sessions:
        h = health(int(s["port"]))
        if h.get("meeting_state") in ("waiting", "in_meeting", "joining"):
            started.add(s["id"])
            log(f"already live {s['id']} :{s['port']} state={h.get('meeting_state')}")

    while True:
        n = now()
        # Render worker is intentionally persistent.
        # Keep watching for future sessions instead of exiting at 23:45.

        # Periodically pick up join URLs dropped into Excel / .env
        if int(time.time()) % 60 < TICK_SEC + 1:
            try_fill_missing_join_from_excel(sessions)
            # Re-apply force_disabled / enabled from disk, and adopt sessions
            # registered after this runner started (e.g. via register_session.py
            # or the dashboard Connect form) so a long-lived `npm run build`
            # process doesn't need a restart to pick up new webinars.
            try:
                disk = load_manifest_sessions()
                by_id = {str(x.get("id")): x for x in disk}
                known_ids = {s["id"] for s in sessions}
                for d in disk:
                    if str(d.get("id")) in known_ids:
                        continue
                    sessions.append(d)
                    lead = int(d.get("start_lead_minutes") or 30)
                    st = parse_hhmmss(d["session_start_ist"])
                    go = st - timedelta(minutes=lead)
                    log(
                        f"discovered new session {d['id']} :{d['port']} "
                        f"start={d['session_start_ist']} auto_start={go.strftime('%H:%M:%S')}"
                    )
                    fp = first_payment_from_schedule(ROOT / d["schedule_file"])
                    if fp:
                        first_pay[d["id"]] = fp
                        log(f"  first_payment {d['id']} @ {fp['time']} → {fp['url'][:60]}")
                    h = health(int(d["port"]))
                    if h.get("meeting_state") in ("waiting", "in_meeting", "joining"):
                        started.add(d["id"])
                        log(f"already live {d['id']} :{d['port']} state={h.get('meeting_state')}")
                for s in sessions:
                    d = by_id.get(s["id"])
                    if not d:
                        continue
                    if d.get("force_disabled"):
                        s["force_disabled"] = True
                        s["enabled"] = False
                    elif "enabled" in d:
                        s["enabled"] = bool(d["enabled"])
            except Exception as exc:
                log(f"manifest discovery failed path={MANIFEST}: {type(exc).__name__}: {exc}")
        for s in sessions:
            if s.get("force_disabled"):
                s["enabled"] = False
                continue
            if s.get("enabled") and s.get("join_url_present"):
                continue
            env = ROOT / s["env_file"]
            if not env.exists():
                continue
            for line in env.read_text().splitlines():
                if line.startswith("MEETING_JOIN_URL="):
                    val = line.split("=", 1)[1].strip().strip("'\"")
                    if val.startswith("http"):
                        s["join_url_present"] = True
                        if not s.get("force_disabled"):
                            s["enabled"] = True
                            log(f"enabled {s['id']} — join URL appeared in {s['env_file']}")
                    break

        for s in sessions:
            if s.get("force_disabled") or not s.get("enabled"):
                continue
            sid = s["id"]
            port = int(s["port"])

            # session_start_ist/session_end_ist are time-of-day only, so a session
            # registered on a previous day looks identical to one registered today
            # once the clock rolls past midnight — without this check it would be
            # auto-(re)started at the same clock time forever. WEBINAR_END_AT in the
            # env file carries the real date; a generous 3h grace keeps late sheet
            # syncs/retries (which already run up to end+120min) unaffected.
            real_end = env_datetime(s["env_file"], "WEBINAR_END_AT")
            if real_end and now() > real_end + timedelta(hours=3):
                if sid in started or health(port).get("meeting_state") in ("waiting", "in_meeting", "joining"):
                    log(f"expiring stale session {sid} :{port} (ended {real_end.isoformat()}) — stopping bridge")
                    subprocess.run([str(HERMES), "stop", str(port)], cwd=str(ROOT), check=False)
                    started.discard(sid)
                s["force_disabled"] = True
                s["enabled"] = False
                persist_force_disable(sid)
                continue

            lead = int(s.get("start_lead_minutes") or 30)
            peak_mins = int(s.get("peak_window_minutes") or 60)
            start_dt = parse_hhmmss(s["session_start_ist"])
            persistent = bool(s.get("keep_connected")) or not s.get("session_end_ist")
            end_dt = (
                parse_hhmmss(s["session_end_ist"])
                if not persistent
                else n + timedelta(days=36500)
            )
            auto_at = start_dt - timedelta(minutes=lead)
            peak_until = start_dt + timedelta(minutes=peak_mins)

            # Auto-start window: from lead time until session end
            if sid not in started and n >= auto_at and n <= end_dt + timedelta(minutes=10):
                if start_slot(port, s["env_file"]):
                    started.add(sid)

            # Peak tracking: session start → +1h, only if bridge is live
            if start_dt <= n <= peak_until:
                last = last_peak_at.get(sid, 0.0)
                if time.time() - last >= PEAK_INTERVAL:
                    last_peak_at[sid] = time.time()
                    h = health(port)
                    state = h.get("meeting_state") or "down"
                    if state in ("in_meeting", "waiting"):
                        c = eval_counts(port)
                        key = f"{n.strftime('%Y-%m-%d')}:{port}:{sid}"
                        slot = peak["slots"].setdefault(
                            key,
                            {
                                "port": port,
                                "id": sid,
                                "meeting_id": s["meeting_id"],
                                "session_start_ist": s["session_start_ist"],
                                "date": n.strftime("%Y-%m-%d"),
                                "peak_footer": 0,
                                "peak_attendees": 0,
                                "peak_participants": 0,
                                "last": {},
                            },
                        )
                        for field, peak_key in (
                            ("footer", "peak_footer"),
                            ("attendees", "peak_attendees"),
                            ("participants", "peak_participants"),
                        ):
                            v = c.get(field)
                            if isinstance(v, int) and v > (slot.get(peak_key) or 0):
                                slot[peak_key] = v
                                slot[f"{peak_key}_at"] = n.isoformat()
                        slot["last"] = {
                            **c,
                            "meeting_state": state,
                            "at": n.isoformat(),
                        }
                        peak["updated_at"] = n.isoformat()
                        save_peak(peak)
                        log(
                            f"peak {sid} :{port} state={state} now={c} "
                            f"peaks=f:{slot['peak_footer']} a:{slot['peak_attendees']} p:{slot['peak_participants']}"
                        )

            # First payment-link drop → snapshot attendees once (per calendar day)
            fp = first_pay.get(sid)
            day_key = f"{n.strftime('%Y-%m-%d')}:{sid}"
            existing = (pay.get("sessions") or {}).get(day_key)
            if fp and not (existing and existing.get("attendees_count") is not None):
                drop_at = parse_hhmmss(fp["time"])
                # Capture in a short window around the scheduled drop (±90s)
                if abs((n - drop_at).total_seconds()) <= 90:
                    h = health(port)
                    state = h.get("meeting_state") or "down"
                    if state in ("in_meeting", "waiting"):
                        c = eval_counts(port)
                        capture_payment_drop(s, fp, c, pay)

            # End-of-session → deliver to Slack, then forget all runtime state.
            sync_at = end_dt + timedelta(minutes=SHEET_SYNC_GRACE_MIN)
            report_key = f"{n.strftime('%Y-%m-%d')}:{sid}"
            report_status = (slack_reports.get("sessions") or {}).get(report_key) or {}
            if (
                sid in started
                and not report_status.get("completed")
                and n >= sync_at
                and n <= end_dt + timedelta(minutes=120)
            ):
                peak_row = metric_row(peak, s.get("meeting_id"), payment=False)
                payment_row = metric_row(pay, s.get("meeting_id"), payment=True)
                try:
                    result = finalize_report(s, peak_row, payment_row, {})
                    delivery = result.get("delivery") or {}
                    if delivery.get("ok"):
                        log(f"slack_report OK {sid} :{port}")
                        subprocess.run([str(HERMES), "stop", str(port)], cwd=str(ROOT), check=False)
                        forget_session(s, result.get("report_files") or [])
                        started.discard(sid)
                        sessions[:] = [item for item in sessions if str(item.get("id")) != sid]
                        first_pay.pop(sid, None)
                        slack_reports.setdefault("sessions", {}).pop(report_key, None)
                        save_slack_reports(slack_reports)
                    else:
                        report_status = {
                            "completed": False,
                            "date": n.strftime("%Y-%m-%d"),
                            "slack_ok": False,
                            "slack_error": delivery.get("error") or delivery.get("skipped"),
                            "last_attempt_at": n.isoformat(),
                        }
                        slack_reports.setdefault("sessions", {})[report_key] = report_status
                        save_slack_reports(slack_reports)
                        log(f"slack_report skipped/failed {sid} :{port} — retaining state for retry")
                except Exception as e:
                    log(f"slack_report failed {sid} :{port} — finalization continued: {type(e).__name__}")

        # Refresh docs/DAILY-RUN-TODAY.md periodically
        if time.time() - last_daily_md >= DAILY_MD_INTERVAL:
            last_daily_md = time.time()
            try:
                r = subprocess.run(
                    [sys.executable, str(ROOT / "scripts/write_daily_run.py")],
                    cwd=str(ROOT),
                    capture_output=True,
                    text=True,
                    timeout=60,
                )
                if r.returncode == 0:
                    log("daily_run.md refreshed")
                else:
                    log(f"daily_run.md refresh failed: {r.stderr[-200:]}")
            except Exception as e:
                log(f"daily_run.md refresh error: {e}")

        time.sleep(TICK_SEC)


if __name__ == "__main__":
    main()
