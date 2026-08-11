"""Parse uploaded CSV/XLSX sheets into Hermes schedule.json items."""
from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any, Optional

_TIME_HEADER = re.compile(r"^(time|timestamp|when|clock|send_at|send at)$", re.I)
_TEXT_HEADER = re.compile(r"^(text|message|chat|msg|body|content)$", re.I)
_POLL_HEADER = re.compile(r"^(poll|poll_name|poll name)$", re.I)
_POLL_END_HEADER = re.compile(r"^(poll_end|poll end|end_poll)$", re.I)


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, h in enumerate(headers):
        n = _norm_header(h)
        if _TIME_HEADER.match(n) and "time" not in mapping:
            mapping["time"] = i
        elif _TEXT_HEADER.match(n) and "text" not in mapping:
            mapping["text"] = i
        elif _POLL_HEADER.match(n) and "poll" not in mapping:
            mapping["poll"] = i
        elif _POLL_END_HEADER.match(n) and "poll_end" not in mapping:
            mapping["poll_end"] = i
    return mapping


def _excel_serial_to_hms(n: float) -> str:
    # Excel time is fraction of a day
    total = int(round(float(n) * 24 * 3600)) % (24 * 3600)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def normalize_time(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%H:%M:%S")
    if isinstance(raw, time):
        return raw.strftime("%H:%M:%S")
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        # Excel serial time (0–1) or day+time
        n = float(raw)
        if 0 <= n < 1.5:
            return _excel_serial_to_hms(n % 1)
        # seconds since midnight
        if n < 86400:
            total = int(n)
            h, rem = divmod(total, 3600)
            m, s = divmod(rem, 60)
            return f"{h:02d}:{m:02d}:{s:02d}"

    s = str(raw).strip()
    if not s:
        return None

    # 8:05pm / 8:05 am
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?\s*([ap]m)$", s, re.I)
    if m:
        h = int(m.group(1))
        mi = int(m.group(2))
        sec = int(m.group(3) or 0)
        ap = m.group(4).lower()
        if ap == "pm" and h != 12:
            h += 12
        if ap == "am" and h == 12:
            h = 0
        return f"{h:02d}:{mi:02d}:{sec:02d}"

    # HH:MM[:SS]
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if m:
        h, mi, sec = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        if h > 23 or mi > 59 or sec > 59:
            return None
        return f"{h:02d}:{mi:02d}:{sec:02d}"

    # ISO datetime → clock
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%H:%M:%S")
    except ValueError:
        pass

    return None


def rows_to_items(rows: list[list[Any]], headers: list[str]) -> list[dict[str, str]]:
    mapping = _map_headers(headers)
    if "time" not in mapping:
        # Assume col0=time, col1=text if no headers match
        if len(headers) >= 2 and not any(headers):
            mapping = {"time": 0, "text": 1}
        elif len(headers) >= 2 and all(not _TIME_HEADER.match(_norm_header(h)) for h in headers):
            # first row might be data — treat positional
            mapping = {"time": 0, "text": 1}
            if len(headers) > 2:
                mapping["poll"] = 2
        else:
            raise ValueError(
                "Sheet needs a Time column (and Text and/or Poll). "
                "Headers like: time, text, poll"
            )
    if "text" not in mapping and "poll" not in mapping:
        raise ValueError("Sheet needs a Text and/or Poll column")

    items: list[dict[str, str]] = []
    for row in rows:
        if not row or all(str(c or "").strip() == "" for c in row):
            continue
        def cell(key: str) -> str:
            idx = mapping.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        when = normalize_time(row[mapping["time"]] if mapping["time"] < len(row) else None)
        if not when:
            continue
        text = cell("text")
        poll = cell("poll")
        poll_end = cell("poll_end")
        if not text and not poll:
            continue
        item: dict[str, str] = {"time": when}
        if poll:
            item["poll"] = poll
        if poll_end:
            item["poll_end"] = poll_end
        if text:
            item["text"] = text
        items.append(item)

    if not items:
        raise ValueError("No valid schedule rows found (need time + text/poll)")
    items.sort(key=lambda x: x["time"])
    return items


def parse_csv_bytes(data: bytes, filename: str = "") -> list[dict[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    # sniff delimiter
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = list(reader)
    if not all_rows:
        raise ValueError("CSV is empty")
    headers = [str(h) for h in all_rows[0]]
    # If first row looks like a time, treat as data with positional cols
    if normalize_time(headers[0]):
        return rows_to_items(all_rows, ["time", "text", "poll"][: len(headers)])
    return rows_to_items(all_rows[1:], headers)


def parse_xlsx_bytes(data: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "XLSX support needs openpyxl. Run: pip install openpyxl "
            "(or upload CSV instead)"
        ) from e
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Spreadsheet is empty") from e
    headers = ["" if c is None else str(c) for c in first]
    body: list[list[Any]] = [list(r) for r in rows_iter]
    if normalize_time(headers[0] if headers else None):
        return rows_to_items([list(first), *body], ["time", "text", "poll"][: len(headers)])
    return rows_to_items(body, headers)


def parse_upload(filename: str, data: bytes) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return parse_csv_bytes(data, filename)
    # sniff
    if data[:2] == b"PK":
        return parse_xlsx_bytes(data)
    return parse_csv_bytes(data, filename)


def write_schedule(
    schedule_path: Path,
    *,
    meeting_id: str,
    session_name: str,
    items: list[dict[str, str]],
    session_start_ist: str = "",
    session_end_ist: str = "",
) -> dict[str, Any]:
    """Write session-scoped schedule JSON. Preserves metadata when present."""
    existing: dict[str, Any] = {}
    if schedule_path.exists():
        try:
            raw = json.loads(schedule_path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                existing = raw
        except json.JSONDecodeError:
            existing = {}

    start = session_start_ist or existing.get("session_start_ist") or ""
    end = session_end_ist or existing.get("session_end_ist") or ""
    if not start and items:
        start = items[0]["time"]
    if not end and items:
        end = items[-1]["time"]

    out = {
        "meeting_id": str(meeting_id),
        "session": session_name or existing.get("session") or schedule_path.stem,
        "session_start_ist": start,
        "session_end_ist": end,
        "items": items,
    }
    schedule_path.parent.mkdir(parents=True, exist_ok=True)
    schedule_path.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return out
